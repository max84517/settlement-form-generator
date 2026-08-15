"""
Generate Word settlement contracts from merged data.

Key responsibilities:
  1. Group data by contract (supplier key + consistent ICM header fields)
  2. Replace plain-text keywords in the Word template
  3. Populate the platform/payment table
  4. Save one .docx per contract group
  5. Optionally update Status in the source Excel to "Contract Generated"
"""
from __future__ import annotations

import copy
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from docx import Document
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml.ns import qn
from docx.oxml import OxmlElement

from settlement_form.utils.amount_utils import to_capital_words

# The 12 ICM header fields that must be consistent within one contract
_ICM_FIELDS = [
    "ICMAgreementCode",
    "ICMPartyName1",
    "ICMSRAGREEMENTCODE",
    "ICMSRAgreementEffectiveDate",
    "ICMEffectiveDate",
    "GTK Supplier",
    "Sub-Category",
    "ICMInternalSignatory",
    "ICMInternalSignatoryTitle",
    "ICMExternalSignatory",
    "ICMExternalSignatoryTitle",
    "CW#",
]

# Map template keywords → DataFrame column names
_KEYWORD_MAP: dict[str, str] = {
    "ICMAgreementCode":          "ICMAgreementCode",
    "ICMPartyName1":             "ICMPartyName1",
    "ICMSRAgreementCode":        "ICMSRAGREEMENTCODE",
    "ICMSRAgreementEffectiveDate": "ICMSRAgreementEffectiveDate",
    "ICMEffectiveDate":          "ICMEffectiveDate",
    "GTK Supplier":              "GTK Supplier",
    "Sub-Category":              "Sub-Category",
    "ICMInternalSignatory":      "ICMInternalSignatory",
    "ICMInternalSignatoryTitle": "ICMInternalSignatoryTitle",
    "ICMExternalSignatory":      "ICMExternalSignatory",
    "ICMExternalSignatoryTitle": "ICMExternalSignatoryTitle",
    "CW#":                       "CW#",
}


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------

def group_by_contract(df: pd.DataFrame) -> list[pd.DataFrame]:
    """
    Split df into contract groups.

    Each group shares the same _supplier_key AND the same values for all
    _ICM_FIELDS (after filling NaN with "").  Different value combinations
    within the same supplier produce separate contracts.
    """
    # Fill NaN in ICM fields so groupby works cleanly
    icm_present = [c for c in _ICM_FIELDS if c in df.columns]
    work = df.copy()
    for c in icm_present:
        work[c] = work[c].fillna("").astype(str)

    group_cols = ["_supplier_key"] + icm_present
    groups: list[pd.DataFrame] = []
    for _, g in work.groupby(group_cols, dropna=False, sort=False):
        groups.append(g.reset_index(drop=True))
    return groups


# ---------------------------------------------------------------------------
# Keyword replacement (plain-text, no delimiters)
# ---------------------------------------------------------------------------

_XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"


def _run_text(run_elem: Any) -> str:
    """Return the combined text of all <w:t> elements inside a <w:r>."""
    return "".join(t.text or "" for t in run_elem.findall(qn("w:t")))


def _set_run_text(run_elem: Any, text: str) -> None:
    """Write `text` into the first <w:t> of a run; blank out the rest."""
    t_elems = run_elem.findall(qn("w:t"))
    if not t_elems:
        new_t = OxmlElement("w:t")
        run_elem.append(new_t)
        t_elems = [new_t]
    t_elems[0].text = text
    if text != text.strip():
        t_elems[0].set(_XML_SPACE, "preserve")
    elif _XML_SPACE in t_elems[0].attrib:
        del t_elems[0].attrib[_XML_SPACE]
    for t in t_elems[1:]:
        t.text = ""


def _strip_caps_from_run(run_elem: Any) -> None:
    """
    Remove <w:caps/> and <w:smallCaps/> from a run's properties so that
    the replacement value is displayed in its own case (not forced to ALL CAPS).
    """
    rpr = run_elem.find(qn("w:rPr"))
    if rpr is None:
        return
    for tag in (qn("w:caps"), qn("w:smallCaps")):
        for elem in rpr.findall(tag):
            rpr.remove(elem)


def _replace_in_p_element(p_elem: Any, replacements: dict[str, str]) -> None:
    """
    Replace keywords inside a single <w:p> XML element.

    Strategy (preserves per-run formatting like bold/italic):
      1. Sort keywords longest-first so e.g. ICMInternalSignatoryTitle is
         matched before its substring ICMInternalSignatory.
      2. For each keyword, try to find it entirely within ONE run and
         replace it there — the run keeps its original formatting (bold, etc.)
      3. Only if the keyword spans multiple runs fall back to merging all run
         texts into the first run (run[0]'s formatting wins, unavoidable).
    """
    runs = [
        r for r in p_elem.findall(".//" + qn("w:r"))
        if r.find(qn("w:t")) is not None
    ]
    if not runs:
        return

    # Longest keywords first → prevents substring substitution
    sorted_kw = sorted(replacements.items(), key=lambda x: -len(x[0]))

    for keyword, value in sorted_kw:
        # Quick check against current joined text (reflects earlier replacements)
        full_text = "".join(_run_text(r) for r in runs)
        if not re.search(re.escape(keyword), full_text, re.IGNORECASE):
            continue

        # ── Try single-run replacement (preserves that run's formatting) ──
        replaced_in_single = False
        for run in runs:
            rt = _run_text(run)
            if re.search(re.escape(keyword), rt, re.IGNORECASE):
                new_rt = re.sub(
                    re.escape(keyword),
                    lambda _m, v=value: v,
                    rt,
                    flags=re.IGNORECASE,
                )
                _set_run_text(run, new_rt)
                _strip_caps_from_run(run)
                replaced_in_single = True
                break

        if replaced_in_single:
            continue

        # ── Keyword spans multiple runs: merge into runs[0] ──────────────
        new_full = re.sub(
            re.escape(keyword),
            lambda _m, v=value: v,
            full_text,
            flags=re.IGNORECASE,
        )
        _set_run_text(runs[0], new_full)
        _strip_caps_from_run(runs[0])
        for run in runs[1:]:
            for t in run.findall(qn("w:t")):
                t.text = ""


def replace_keywords_in_doc(doc: Document, replacements: dict[str, str]) -> None:
    """
    Apply keyword replacements everywhere in the document.

    Iterates the ENTIRE XML tree so text boxes, SDT content controls,
    frames, headers, and footers are all covered — not just the paragraphs
    exposed by python-docx's high-level API.
    """
    # All <w:p> elements anywhere in the document body
    for p_elem in doc.element.body.iter(qn("w:p")):
        _replace_in_p_element(p_elem, replacements)

    # Headers and footers in every section
    for section in doc.sections:
        for hf in (
            section.header,
            section.footer,
            section.even_page_header,
            section.even_page_footer,
            section.first_page_header,
            section.first_page_footer,
        ):
            if hf is not None:
                for p_elem in hf._element.iter(qn("w:p")):
                    _replace_in_p_element(p_elem, replacements)


# ---------------------------------------------------------------------------
# Table population
# ---------------------------------------------------------------------------

def _fill_tc(tc_elem: Any, text: str, bold: bool = False,
             align_val: str = "left") -> None:
    """
    Set the content of a table cell <w:tc> element directly via XML.

    Works on cells found anywhere in the XML tree (including inside <w:sdt>
    content controls), bypassing python-docx's high-level Table/Cell API
    which only sees top-level tables.
    """
    # Keep <w:tcPr> (width, borders) but replace the paragraph content.
    p_elem = tc_elem.find(qn("w:p"))
    if p_elem is None:
        p_elem = OxmlElement("w:p")
        tc_elem.append(p_elem)

    # Remove all existing runs (preserve paragraph / cell properties)
    for r in p_elem.findall(".//" + qn("w:r")):
        r.getparent().remove(r)

    # Set paragraph alignment
    ppr = p_elem.find(qn("w:pPr"))
    if ppr is None:
        ppr = OxmlElement("w:pPr")
        p_elem.insert(0, ppr)
    jc = ppr.find(qn("w:jc"))
    if jc is None:
        jc = OxmlElement("w:jc")
        ppr.append(jc)
    jc.set(qn("w:val"), align_val)

    # Build new run
    r_elem = OxmlElement("w:r")
    if bold:
        rpr = OxmlElement("w:rPr")
        rpr.append(OxmlElement("w:b"))
        r_elem.append(rpr)
    t_elem = OxmlElement("w:t")
    t_elem.text = text
    if text and text != text.strip():
        t_elem.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    r_elem.append(t_elem)
    p_elem.append(r_elem)


def fill_platform_table(
    doc: Document,
    platform_rows: list[tuple[str, str, float]],
    table_index: int = 0,
) -> None:
    """
    Populate the platform/payment table in the document.

    Searches the FULL XML tree for <w:tbl> elements (handles tables inside
    <w:sdt> content controls which doc.tables misses).  Fills pre-existing
    data rows first, appends new rows if needed, clears leftover rows, and
    appends a bold Total row at the end.

    Parameters
    ----------
    platform_rows : list of (sub_category, platform, amount)
    table_index   : index into the list of ALL tables in the XML tree
    """
    # Find all tables anywhere in the document body (including inside SDTs)
    all_tbl_elems = doc.element.body.findall(".//" + qn("w:tbl"))
    if not all_tbl_elems or table_index >= len(all_tbl_elems):
        return

    tbl_elem = all_tbl_elems[table_index]
    total = sum(amount for _, _, amount in platform_rows)

    all_tr = tbl_elem.findall(qn("w:tr"))       # direct-child rows only
    if not all_tr:
        return

    existing_data_tr = all_tr[1:]               # skip header row (index 0)

    for i, (sub_cat, platform, amount) in enumerate(platform_rows):
        if i < len(existing_data_tr):
            tr_elem = existing_data_tr[i]       # reuse pre-existing row
        else:
            # Clone the last available data row (preserves borders/shading)
            src = existing_data_tr[-1] if existing_data_tr else all_tr[0]
            tr_elem = copy.deepcopy(src)
            tbl_elem.append(tr_elem)

        tc_elems = tr_elem.findall(qn("w:tc"))
        if len(tc_elems) >= 2:
            _fill_tc(tc_elems[0], f"[{sub_cat}/{platform}] E&O GTK Parts", align_val="left")
            _fill_tc(tc_elems[1], f"${amount:,.2f}", align_val="left")

    # Clear any leftover pre-existing rows that received no data
    for j in range(len(platform_rows), len(existing_data_tr)):
        for tc in existing_data_tr[j].findall(qn("w:tc")):
            _fill_tc(tc, "", align_val="left")

    # Total row — always a fresh row appended at the bottom
    src = existing_data_tr[-1] if existing_data_tr else all_tr[0]
    total_tr = copy.deepcopy(src)
    tbl_elem.append(total_tr)
    tc_elems = total_tr.findall(qn("w:tc"))
    if len(tc_elems) >= 2:
        _fill_tc(tc_elems[0], "Total Settlement Payment (in USD)",
                 bold=True, align_val="right")
        _fill_tc(tc_elems[1], f"${total:,.2f}", align_val="left")


# ---------------------------------------------------------------------------
# Contract generation
# ---------------------------------------------------------------------------

def _to_date_str(val: Any) -> str:
    """
    Convert a date value (datetime, pandas Timestamp, or string) to
    'MMM D, YYYY' format, e.g. 'May 1, 2026' or 'Nov 13, 2024'.
    Returns the original string unchanged if parsing fails.
    """
    from datetime import datetime, date as _date
    import pandas as _pd

    # pd.isna() handles NaN, NaT, and None; wrap in try for exotic types
    try:
        if _pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass

    if isinstance(val, str) and not val.strip():
        return ""

    dt = None
    if isinstance(val, (_date, datetime)):
        dt = val
    else:
        try:
            ts = _pd.Timestamp(val)
            # pd.NaT.to_pydatetime() doesn't raise but returns NaT, which
            # then fails on .strftime() — check isna() before converting.
            if _pd.isna(ts):
                return ""
            dt = ts.to_pydatetime()
        except Exception:
            pass

    if dt is None:
        return str(val).strip()

    return f"{dt.strftime('%B')} {dt.day}, {dt.strftime('%Y')}"


def _build_replacements(group: pd.DataFrame) -> dict[str, str]:
    """Build the keyword→value mapping for a single contract group."""
    first = group.iloc[0]
    replacements: dict[str, str] = {}

    for keyword, col in _KEYWORD_MAP.items():
        if col in group.columns:
            raw = first[col]
            if not pd.notna(raw):
                replacements[keyword] = ""
            elif col == "ICMSRAgreementEffectiveDate":
                # Format as 'MMM D, YYYY'
                replacements[keyword] = _to_date_str(raw)
            elif col == "ICMAgreementCode":
                # First letter uppercase, rest lowercase
                replacements[keyword] = str(raw).capitalize()
            else:
                replacements[keyword] = str(raw)

    # Computed keywords
    platform_rows = _get_platform_rows(group)
    total = sum(amt for _, _, amt in platform_rows)

    replacements["TOTALPAYMENT"] = f"${total:,.2f}"
    replacements["CAPITALLETTERSAMOUNT"] = to_capital_words(total)

    platform_list = ", ".join(p for _, p, _ in platform_rows)
    replacements["PLATFORMLIST"] = platform_list

    return replacements


def _get_platform_rows(group: pd.DataFrame) -> list[tuple[str, str, float]]:
    """Return (sub_category, platform, amount) tuples sorted by platform."""
    rows = []
    for _, r in group.iterrows():
        sub_cat = str(r.get("Sub-Category", "")).strip()
        platform = str(r.get("Platform", "")).strip()
        amount = float(r.get("Actual Payment", 0.0))
        rows.append((sub_cat, platform, amount))
    return sorted(rows, key=lambda x: x[1])


def _make_output_dir(output_folder: str | Path) -> Path:
    """Create and return a timestamped output directory."""
    ts = datetime.now().strftime("%Y-%m-%d %H-%M")
    out_dir = Path(output_folder) / ts
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _safe_filename(name: str) -> str:
    """Remove characters that are invalid in Windows filenames."""
    return re.sub(r'[\\/:*?"<>|]', "_", name)


# Fields that must be non-empty for a complete contract
_REQUIRED_FIELDS = [
    "ICMAgreementCode",
    "ICMPartyName1",
    "ICMSRAGREEMENTCODE",
    "ICMSRAgreementEffectiveDate",
    "ICMEffectiveDate",
    "ICMInternalSignatory",
    "ICMInternalSignatoryTitle",
    "ICMExternalSignatory",
    "ICMExternalSignatoryTitle",
    "CW#",
]


def check_missing_fields(df: pd.DataFrame) -> dict[str, list[str]]:
    """
    Return a dict mapping supplier_key → list of empty required fields.
    Only suppliers that have at least one missing field are included.
    An empty dict means all data is complete.
    """
    result: dict[str, list[str]] = {}
    for supplier_key, group in df.groupby("_supplier_key", sort=False):
        first = group.iloc[0]
        missing = [
            col for col in _REQUIRED_FIELDS
            if col in df.columns and (
                pd.isna(first[col]) or str(first[col]).strip() == ""
            )
        ]
        if missing:
            result[str(supplier_key)] = missing
    return result


def generate_contracts(
    df: pd.DataFrame,
    template_path: str | Path,
    output_folder: str | Path,
) -> list[Path]:
    """
    Generate one .docx contract per group.

    Returns the list of output file paths.
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    out_dir = _make_output_dir(output_folder)
    groups = group_by_contract(df)
    output_paths: list[Path] = []

    # Count groups per (supplier_key, icm_code) to detect multi-Sub-Category suppliers
    from collections import Counter
    supplier_code_count: Counter = Counter()
    for g in groups:
        f = g.iloc[0]
        supplier_code_count[(str(f.get("_supplier_key", "")),
                              str(f.get("ICMAgreementCode", "")).strip())] += 1

    for group in groups:
        first = group.iloc[0]
        supplier_key = str(first.get("_supplier_key", "Unknown"))
        icm_code = str(first.get("ICMAgreementCode", "")).strip()
        sub_cat = str(first.get("Sub-Category", "")).strip()

        doc = Document(template_path)
        replacements = _build_replacements(group)
        replace_keywords_in_doc(doc, replacements)

        platform_rows = _get_platform_rows(group)
        fill_platform_table(doc, platform_rows)

        safe_supplier = _safe_filename(supplier_key)
        safe_code = _safe_filename(icm_code) if icm_code else "NO_CODE"
        # Include Sub-Category in filename only when the supplier has multiple Sub-Categories
        if supplier_code_count[(supplier_key, icm_code)] > 1:
            filename = f"SETTLEMENT AND RELEASE AGREEMENT_{safe_supplier}_{safe_code}_{_safe_filename(sub_cat)}.docx"
        else:
            filename = f"SETTLEMENT AND RELEASE AGREEMENT_{safe_supplier}_{safe_code}.docx"
        out_path = out_dir / filename
        doc.save(out_path)
        output_paths.append(out_path)

    return output_paths


# ---------------------------------------------------------------------------
# Status update
# ---------------------------------------------------------------------------

def update_status_in_excel(
    excel_path: str | Path,
    original_df: pd.DataFrame,
    processed_df: pd.DataFrame,
) -> None:
    """
    Update the 'Status' column to 'Contract Generated' for rows in
    processed_df, matched by (GTK Supplier, Platform, Sub-Category).

    We load the workbook directly to preserve formatting, formulas, etc.
    """
    excel_path = Path(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    # Find the header row (row 1) and locate the Status column index
    header = {
        str(cell.value).replace("\n", " ").strip(): cell.column
        for cell in ws[1]
        if cell.value
    }
    status_col = header.get("Status")
    supplier_col = header.get("GTK Supplier")
    platform_col = header.get("Platform")
    subcat_col = header.get("Sub-Category")

    if not all([status_col, supplier_col, platform_col, subcat_col]):
        raise ValueError("Could not find required columns in source Excel.")

    def _norm(s: str) -> str:
        return s.lower().replace(" ", "")

    processed_set: set[tuple[str, str, str]] = set()
    for _, row in processed_df.iterrows():
        processed_set.add((
            _norm(str(row.get("GTK Supplier", ""))),
            _norm(str(row.get("Platform", ""))),
            _norm(str(row.get("Sub-Category", ""))),
        ))

    for row in ws.iter_rows(min_row=2):
        supplier = _norm(str(row[supplier_col - 1].value or ""))
        platform = _norm(str(row[platform_col - 1].value or ""))
        subcat   = _norm(str(row[subcat_col - 1].value or ""))
        if (supplier, platform, subcat) in processed_set:
            row[status_col - 1].value = "Contract Generated"

    wb.save(excel_path)
